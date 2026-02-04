import os
import sys
import io
import pandas as pd
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ---------------------------
# Load environment variables
# ---------------------------
load_dotenv()

# ---------------------------
# Configuration
# ---------------------------
INPUT_FILE = "results/testcase.txt"
AGENDA_FILE = "agend.md"
REQUIREMENT_FILE = "resource/requirement.md"
OUTPUT_FILE = "results/testcase.xlsx"
MODEL_NAME = "gpt-4.1"

EXPECTED_COLUMNS = [
    "TC ID",
    "Test Case Name",
    "Test Objective",
    "Prepare Step",
    "Test Data",
    "Test Steps",
    "Expected Result"
]

# ---------------------------
# Utility Functions
# ---------------------------

def clean_ai_output(text: str) -> str:
    """
    Remove markdown code fences such as ``` or ```tsv from AI output.
    """
    text = text.strip()
    if text.startswith("```"):
        parts = text.split("```", 2)
        if len(parts) >= 2:
            text = parts[1]
    if text.endswith("```"):
        text = text.rsplit("```", 1)[0]
    return text.strip()


def normalize_multiline_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize multiline columns:
    - Replace structural delimiter ( | ) with newline
    - Remove leading spaces after newline
    - Preserve real '|' inside values
    """
    multiline_columns = ["Test Data", "Test Steps"]

    for col in multiline_columns:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                # replace ONLY delimiter-style pipes
                .str.replace(r"\s+\|\s+", "\n", regex=True)
                .str.replace(r"\n\s+", "\n", regex=True)
                .str.strip()
            )

    return df


def normalize_prepare_step(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare Step is OPTIONAL.
    Remove meaningless values and keep only real preparation steps.
    """
    meaningless_values = {
        "n/a",
        "na",
        "none",
        "no",
        "no preparation required",
        "not required",
        "-",
        ""
    }

    def clean_prepare(val):
        if not isinstance(val, str):
            return ""
        normalized = val.strip().lower()
        if normalized in meaningless_values:
            return ""
        return val.strip()

    if "Prepare Step" in df.columns:
        df["Prepare Step"] = df["Prepare Step"].apply(clean_prepare)

    return df


def validate_output(df: pd.DataFrame):
    """
    Validate AI output to ensure it is Sheet-safe and requirement-aligned.
    """
    # 1. Column validation
    if list(df.columns) != EXPECTED_COLUMNS:
        raise ValueError(
            f"Column mismatch.\nExpected: {EXPECTED_COLUMNS}\nActual: {list(df.columns)}"
        )

    # 2. API testcase must not contain UI wording
    ui_keywords = ["page", "screen", "click", "form", "button", "navigate"]
    for idx, row in df.iterrows():
        steps = str(row["Test Steps"]).lower()
        if any(word in steps for word in ui_keywords):
            raise ValueError(
                f"UI wording detected in Test Steps at row {idx + 1}: {row['Test Steps']}"
            )


def post_process_excel(file_path: str):
    """
    Enable wrap text and auto row height for cells containing newline.
    Prevent Excel / Google Sheets newline rendering issues.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "\n" in cell.value:
                cell.alignment = wrap_alignment

    # Auto-adjust row height
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = None

    wb.save(file_path)

# ---------------------------
# Main Process
# ---------------------------

def main():
    # 1. Check API Key
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("❌ OPENAI_API_KEY not set.")
        sys.exit(1)

    client = OpenAI(api_key=api_key)

    # 2. Check input files
    for file_path in [INPUT_FILE, AGENDA_FILE]:
        if not os.path.exists(file_path):
            print(f"❌ Required file not found: {file_path}")
            sys.exit(1)

    print("📖 Reading input files...")

    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        testcase_content = f.read()

    with open(AGENDA_FILE, "r", encoding="utf-8") as f:
        agenda_content = f.read()

    requirement_content = ""
    if REQUIREMENT_FILE and os.path.exists(REQUIREMENT_FILE):
        with open(REQUIREMENT_FILE, "r", encoding="utf-8") as f:
            requirement_content = f.read().strip()
            if requirement_content:
                print(f"  ✓ Loaded requirements from {REQUIREMENT_FILE}")

    # 3. Build system prompt
    system_content = f"""
CRITICAL RULES (NON-NEGOTIABLE):

1. You MUST read and follow resource/requirement.md first.
2. requirement.md is the single source of truth.
3. API / Database testcases MUST NOT contain UI wording.
4. Prepare Step is OPTIONAL:
   - If no preparation is required, leave the cell EMPTY.
   - Do NOT use N/A, None, -, or placeholders.
"""

    system_content += "\n\n---\n\n# AGENDA\n\n" + agenda_content

    if requirement_content:
        system_content += "\n\n---\n\n# REQUIREMENTS\n\n" + requirement_content

    # 4. Call OpenAI
    print(f"🤖 Sending request to OpenAI ({MODEL_NAME})...")
    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": system_content},
            {"role": "user", "content": testcase_content}
        ],
        temperature=0.1
    )

    ai_output = response.choices[0].message.content

    # 5. Process AI output
    print("🧹 Cleaning AI output...")
    cleaned_output = clean_ai_output(ai_output)

    # 6. Convert to DataFrame
    try:
        lines = cleaned_output.splitlines()
        first_row = lines[0].split("\t")

        if first_row == EXPECTED_COLUMNS:
            df = pd.read_csv(io.StringIO(cleaned_output), sep="\t")
        else:
            df = pd.read_csv(
                io.StringIO(cleaned_output),
                sep="\t",
                header=None,
                names=EXPECTED_COLUMNS
            )

        df = normalize_multiline_columns(df)
        df = normalize_prepare_step(df)
        validate_output(df)

    except Exception as e:
        print("❌ Validation failed:")
        print(e)
        print("\n--- RAW AI OUTPUT ---\n")
        print(cleaned_output)
        sys.exit(1)

    # 7. Save to Excel
    df.to_excel(OUTPUT_FILE, index=False)
    post_process_excel(OUTPUT_FILE)

    print(f"✅ Success! Test cases saved to '{OUTPUT_FILE}'.")


if __name__ == "__main__":
    main()
