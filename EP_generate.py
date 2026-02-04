from openpyxl import load_workbook
from collections import OrderedDict
import json

FILE_PATH = "resource/EP_table.xlsx"
OUTPUT_HTML = "results/ep_matrix.html"

# =========================================================
# STEP 1: Read conditions from Excel
# =========================================================
def read_conditions(path):
    wb = load_workbook(path, data_only=True, keep_vba=True)
    ws = wb.active

    conditions = OrderedDict()
    last_condition = None

    for row in ws.iter_rows(min_row=2):
        cond = row[0].value
        v_desc = row[1].value
        v_tag = row[2].value
        x_desc = row[3].value
        x_tag = row[4].value

        if cond:
            last_condition = str(cond).strip()

        if not last_condition:
            continue

        if last_condition not in conditions:
            conditions[last_condition] = {
                "valid": OrderedDict(),
                "invalid": OrderedDict()
            }

        if isinstance(v_tag, str):
            conditions[last_condition]["valid"][v_tag] = v_desc

        if isinstance(x_tag, str):
            conditions[last_condition]["invalid"][x_tag] = x_desc

    return conditions


# =========================================================
# STEP 2: Generate EP Testcases (FIXED EP LOGIC)
# =========================================================
def generate_ep_testcases(conditions):
    conds = list(conditions.keys())

    valid_sets = [list(conditions[c]["valid"].keys()) for c in conds]
    invalid_sets = [list(conditions[c]["invalid"].keys()) for c in conds]

    tcs = []

    # -----------------------------------------------------
    # VALID TESTCASES (index-based alignment, NO duplicate)
    # -----------------------------------------------------
    max_len = max(len(v) for v in valid_sets)

    for i in range(max_len):
        tc = []
        for vs in valid_sets:
            tc.append(vs[i % len(vs)])
        tcs.append(tc)

    valid_tc_count = len(tcs)

    # -----------------------------------------------------
    # INVALID TESTCASES (single X per TC)
    # -----------------------------------------------------
    base_idx = 0
    for ci, invs in enumerate(invalid_sets):
        for x in invs:
            base = tcs[base_idx % valid_tc_count].copy()
            base[ci] = x
            tcs.append(base)
            base_idx += 1

    return tcs


# =========================================================
# STEP 3: Print Testcase Detail (KEEP THIS)
# =========================================================
def print_testcases(tcs, conditions):
    conds = list(conditions.keys())

    for i, tc in enumerate(tcs, start=1):
        print(f"\nTC{i}")
        for idx, tag in enumerate(tc):
            cond = conds[idx]
            desc = (
                conditions[cond]["valid"].get(tag)
                or conditions[cond]["invalid"].get(tag)
            )
            print(f"  - {cond}: {tag} = {desc}")


# =========================================================
# STEP 4: Build EP Matrix (NO PRINT)
# =========================================================
def build_matrix(conditions, tcs):
    tags = []
    tag_to_cond = {}

    for cond, data in conditions.items():
        for t in list(data["valid"].keys()) + list(data["invalid"].keys()):
            tags.append(t)
            tag_to_cond[t] = cond

    matrix = OrderedDict((t, [""] * len(tcs)) for t in tags)

    for i, tc in enumerate(tcs):
        for tag in tc:
            matrix[tag][i] = "X"

    return matrix, tag_to_cond


# =========================================================
# STEP 5: Generate HTML Matrix + TC Detail Panel
# =========================================================
def generate_html(matrix, tag_to_cond, tc_count, tcs, conditions):
    conds = list(conditions.keys())

    tc_detail = OrderedDict()
    for i, tc in enumerate(tcs, start=1):
        rows = []
        for idx, tag in enumerate(tc):
            cond = conds[idx]
            desc = (
                conditions[cond]["valid"].get(tag)
                or conditions[cond]["invalid"].get(tag)
            )
            rows.append({
                "condition": cond,
                "tag": tag,
                "desc": "" if desc is None else str(desc)
            })
        tc_detail[f"TC{i}"] = rows

    tc_detail_json = json.dumps(tc_detail, ensure_ascii=False)

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>EP Matrix</title>

<style>
body {{
  font-family: Arial;
  display: flex;
  gap: 24px;
  padding: 16px;
}}

table {{
  border-collapse: collapse;
  width: max-content;
}}

th, td {{
  border: 1px solid #999;
  padding: 6px 12px;
  white-space: nowrap;
  text-align: center;
  font-size: 13px;
}}

th {{
  background: #f3f3f3;
  cursor: pointer;
}}

.tag {{
  text-align: left;
  font-weight: bold;
  min-width: 160px;
  position: sticky;
  left: 0;
  background: white;
}}

.valid {{ background: #c3e6cb; color: #28a745; }}
.invalid {{ background: #f5c6cb; color: #dc3545; }}

.sep-top td {{ border-top: 3px solid black; }}

.highlight-col {{ background: #fff3cd !important; }}
.highlight-green {{ background: #28a745 !important; color: white; font-weight: bold; }}
.highlight-red {{ background: #dc3545 !important; color: white; font-weight: bold; }}

#tc-detail {{
  min-width: 460px;
  border: 1px solid #ccc;
  padding: 16px;
  background: #fafafa;
  max-height: 80vh;
  overflow: auto;
}}
</style>

<script>
const TC_DETAIL = {tc_detail_json};
let activeCol = null;

function toggleColumn(col) {{
  const table = document.getElementById("ep");

  if (activeCol === col) {{
    clearAll();
    document.getElementById("tc-detail").innerHTML = "";
    activeCol = null;
    return;
  }}

  clearAll();
  activeCol = col;

  for (let r = 1; r < table.rows.length; r++) {{
    const cell = table.rows[r].cells[col];
    if (!cell) continue;
    cell.classList.add("highlight-col");

    if (cell.innerText.trim() === "X") {{
      const tagCell = table.rows[r].cells[0];
      const tag = tagCell.innerText.trim().toLowerCase();

      if (tag.startsWith("v")) {{
        cell.classList.add("highlight-green");
        tagCell.classList.add("highlight-green");
      }} else {{
        cell.classList.add("highlight-red");
        tagCell.classList.add("highlight-red");
      }}
    }}
  }}

  table.rows[0].cells[col].classList.add("highlight-col");
  renderTC(col);
}}

function renderTC(col) {{
  const key = "TC" + col;
  const data = TC_DETAIL[key];
  if (!data) return;

  let html = `<div style="display:flex; justify-content:space-between; align-items:center;">
    <h3>${{key}}</h3>
    <button onclick="copyToClipboard('${{key}}')">Copy</button>
  </div>`;

  let copyText = `${{key}}\\n`;

  data.forEach(i => {{
    const cond = escapeHtml(i.condition);
    const tag = escapeHtml(i.tag);
    const desc = escapeHtml(i.desc);
    
    // Plain text for clipboard
    copyText += `- ${{i.condition}}: ${{i.tag}} = ${{i.desc}}\\n`;

    let color = "";
    if (String(i.tag).toLowerCase().startsWith("v")) {{
      color = "color: #28a745;";
    }} else if (String(i.tag).toLowerCase().startsWith("x")) {{
       color = "color: #dc3545;";
    }}

    html += `
      <div style="margin-bottom:8px;">
        - <span style="${{color}}"><b>${{cond}}</b>: <b>${{tag}}</b></span> = ${{desc}}
      </div>
    `;
  }});
  
  document.getElementById("tc-detail").innerHTML = html;
}}

function copyToClipboard(key) {{
  const data = TC_DETAIL[key];
  if (!data) return;

  let text = `${{key}}\\n`;
  data.forEach(i => {{
    text += `- ${{i.condition}}: ${{i.tag}} = ${{i.desc}}\\n`;
  }});

  navigator.clipboard.writeText(text).then(() => {{
    alert("Copied to clipboard!");
  }}).catch(err => {{
    console.error('Failed to copy: ', err);
  }});
}}

function clearAll() {{
  document.querySelectorAll(
    ".highlight-col,.highlight-green,.highlight-red"
  ).forEach(e => e.classList.remove("highlight-col","highlight-green","highlight-red"));
}}

function escapeHtml(text) {{
  if (text === null || text === undefined) return "";
  return String(text)
    .replace(/&/g, '&amp;')
    .replace(/"/g, '&quot;')
    .replace(/'/g, '&#39;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;');
}}
</script>
</head>

<body>
<div>
<h2>Equivalence Partitioning Matrix</h2>
<table id="ep">
<tr>
<th>EP Tag</th>
"""

    for i in range(tc_count):
        html += f'<th onclick="toggleColumn({i+1})">TC{i+1}</th>'

    html += "</tr>"

    prev_cond = None
    for tag, row in matrix.items():
        cond = tag_to_cond[tag]
        cls = "valid" if tag.startswith("v") else "invalid"
        sep = "sep-top" if prev_cond and cond != prev_cond else ""
        html += f'<tr class="{sep}"><td class="tag {cls}">{tag}</td>'
        for c in row:
            html += f"<td>{c}</td>"
        html += "</tr>"
        prev_cond = cond

    html += """
</table>
</div>
<div id="tc-detail"></div>
</body>
</html>
"""
    return html


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    conditions = read_conditions(FILE_PATH)
    tcs = generate_ep_testcases(conditions)

    # Write testcases to text file
    with open("results/testcase.txt", "w", encoding="utf-8") as f:
        conds = list(conditions.keys())
        for i, tc in enumerate(tcs, start=1):
            if i > 1: f.write("\n")
            f.write(f"TC{i}\n")
            for idx, tag in enumerate(tc):
                cond = conds[idx]
                desc = conditions[cond]["valid"].get(tag) or conditions[cond]["invalid"].get(tag) or ""
                f.write(f"  - {cond}: {tag} = {desc}\n")

    # terminal output (unchanged)
    print_testcases(tcs, conditions)

    matrix, tag_to_cond = build_matrix(conditions, tcs)
    html = generate_html(matrix, tag_to_cond, len(tcs), tcs, conditions)

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n[OK] EP Matrix generated → {OUTPUT_HTML}")
